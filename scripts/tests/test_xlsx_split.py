"""Tests para materialización de valores y split XLSX."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook

from xlsx_split import _formula_references_other_sheet, slugify, split_xlsx


def test_slugify_sanitizes_sheet_names():
    assert slugify(" Ítems / Precios ") == "tems_precios"


def test_formula_references_other_sheet():
    assert _formula_references_other_sheet("=Sheet2!A1", "Sheet1")
    assert not _formula_references_other_sheet("=Sheet1!A1", "Sheet1")


def test_split_xlsx_materializes_values_and_splits(tmp_path: Path):
    src = tmp_path / "presupuesto.xlsx"
    wb = Workbook()
    summary = wb.active
    summary.title = "Resumen"
    detail = wb.create_sheet("Detalle")
    detail["A1"] = 99
    summary["A1"] = "Total"
    summary["B1"] = "=Detalle!A1"
    wb.save(src)

    out_dir = tmp_path / "split"
    manifest = split_xlsx(src, out_dir)

    assert manifest["values_materialized"] is True
    assert manifest["total_sheets"] == 2

    summary_path = Path(manifest["sheets"][0]["output_path"])
    wb_single = load_workbook(summary_path, data_only=False)
    try:
        ws = wb_single[summary.title]
        assert ws["B1"].value != "=Detalle!A1"
    finally:
        del wb_single

    manifest_path = out_dir / "presupuesto_split_manifest.json"
    saved = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert saved["stem"] == "presupuesto"
