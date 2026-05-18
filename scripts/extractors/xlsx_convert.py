#!/usr/bin/env python3
"""
xlsx_convert.py — Paso 1.0.b.3 del workflow.

Dado un XLSX single-tab + JSON de análisis (output de xlsx_sheet_analyze.py),
convierte la hoja a la representación óptima decidida en el análisis.

Salidas posibles:
    - markdown_table:  Markdown con tabla en formato pipe (vía MarkItDown).
    - html_table:      HTML con <table>, <thead>, <tbody>, colspan/rowspan
                       (generado nativo con openpyxl para preservar merged cells).
    - markdown_text:   Markdown con celdas concatenadas como párrafos.
    - empty:           archivo vacío con header indicando que se omitió.

Cada output incluye frontmatter YAML con metadata:
    - documento, sheet_name, representation, source_path, fuente_original (si está en manifest).

Uso:
    python3 xlsx_convert.py INPUT_SINGLE_TAB.xlsx [--analysis ANALYSIS.json] [--output PATH]

Si --analysis no se pasa, lo busca como {input_stem}_analysis.json al lado del input.
Si --output no se pasa, lo escribe junto al input con extensión .md.

Determinístico. No LLM.

Dependencias: openpyxl, markitdown (opcional; solo para markdown_table)
"""
from __future__ import annotations

import argparse
import json
import sys
from html import escape
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except ImportError:
    sys.stderr.write(
        "openpyxl no instalado. Instalar con:\n"
        "  pip install --break-system-packages openpyxl\n"
    )
    sys.exit(1)


def _frontmatter(analysis: dict, source_path: Path) -> str:
    fm = [
        "---",
        "tipo: XLSX_NORMALIZADO_HOJA",
        f"source_path: {source_path}",
        f"sheet_name: {json.dumps(analysis.get('sheet_name', ''), ensure_ascii=False)}",
        f"representation: {analysis['representation']}",
        f"max_row: {analysis['max_row']}",
        f"max_col: {analysis['max_col']}",
        f"merged_count: {analysis['merged_count']}",
        f"non_empty_cells: {analysis['non_empty_cells']}",
        "---",
        "",
    ]
    return "\n".join(fm)


def convert_markdown_table(input_path: Path) -> str:
    """Usa MarkItDown si está disponible; si no, fallback a generación manual."""
    try:
        from markitdown import MarkItDown  # type: ignore
        md = MarkItDown()
        result = md.convert(str(input_path))
        # MarkItDown produce el contenido en .text_content
        return result.text_content
    except ImportError:
        # Fallback: generar manualmente
        wb = load_workbook(filename=str(input_path), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ""
        # Asumir fila 1 como header
        header = [str(c) if c is not None else "" for c in rows[0]]
        lines = ["| " + " | ".join(header) + " |"]
        lines.append("|" + "|".join(["---"] * len(header)) + "|")
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            # Padding a longitud de header
            while len(cells) < len(header):
                cells.append("")
            lines.append("| " + " | ".join(cells) + " |")
        return "\n".join(lines)


def convert_html_table(input_path: Path) -> str:
    """
    Genera HTML <table> con colspan/rowspan correctos a partir de openpyxl.
    Preserva merged cells. Heurística: las primeras filas hasta encontrar
    una fila sin merged cells se tratan como <thead>; el resto, <tbody>.
    """
    wb = load_workbook(filename=str(input_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        return "<!-- hoja vacía -->\n"

    # Mapa de celdas "ancla" (top-left de cada merged range) a (rowspan, colspan).
    # Mapa de celdas "tapadas" (parte de merged pero no ancla) → skip.
    anchor_spans = {}  # (row, col) -> (rowspan, colspan)
    masked_cells = set()  # (row, col) tapadas por merged
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col_r, max_row_r = range_boundaries(str(rng))
        rowspan = max_row_r - min_row + 1
        colspan = max_col_r - min_col + 1
        anchor_spans[(min_row, min_col)] = (rowspan, colspan)
        for r in range(min_row, max_row_r + 1):
            for c in range(min_col, max_col_r + 1):
                if (r, c) != (min_row, min_col):
                    masked_cells.add((r, c))

    # Detectar dónde termina el thead: el primer índice de fila en el que
    # no hay merged ranges que toquen esa fila o más arriba.
    rows_with_merged = set()
    for (anchor_r, _), (rs, _) in anchor_spans.items():
        for rr in range(anchor_r, anchor_r + rs):
            rows_with_merged.add(rr)
    thead_end = 0
    if rows_with_merged:
        thead_end = max(r for r in rows_with_merged if r <= 3)  # max fila ≤3 con merged
        if thead_end < 1:
            thead_end = 1  # al menos fila 1 como header

    # Si no hay merged en filas 1-3, asumir fila 1 como header
    if thead_end == 0:
        thead_end = 1

    lines = ['<table>']
    in_thead = False
    in_tbody = False

    for r in range(1, max_row + 1):
        if r <= thead_end:
            if not in_thead:
                lines.append("  <thead>")
                in_thead = True
        else:
            if in_thead:
                lines.append("  </thead>")
                in_thead = False
            if not in_tbody:
                lines.append("  <tbody>")
                in_tbody = True

        lines.append("    <tr>")
        for c in range(1, max_col + 1):
            if (r, c) in masked_cells:
                continue
            cell = ws.cell(row=r, column=c)
            val = cell.value
            text = "" if val is None else str(val)
            text = escape(text).replace("\n", "<br>")
            tag = "th" if r <= thead_end else "td"
            attrs = ""
            if (r, c) in anchor_spans:
                rs, cs = anchor_spans[(r, c)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            if tag == "th":
                attrs += ' scope="col"' if (rs := anchor_spans.get((r, c), (1, 1))[0]) == 1 else ' scope="colgroup"'
            lines.append(f"      <{tag}{attrs}>{text}</{tag}>")
        lines.append("    </tr>")

    if in_thead:
        lines.append("  </thead>")
    if in_tbody:
        lines.append("  </tbody>")
    lines.append("</table>")
    return "\n".join(lines)


def convert_markdown_text(input_path: Path) -> str:
    """
    Convierte una hoja con texto narrativo a markdown plano:
    cada celda no vacía se vuelca como párrafo, en orden de lectura.
    """
    wb = load_workbook(filename=str(input_path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    paragraphs = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None or cell == "":
                continue
            paragraphs.append(str(cell).strip())
    return "\n\n".join(paragraphs)


def convert(input_path: Path, analysis_path: Path | None, output_path: Path | None) -> dict:
    if analysis_path is None:
        analysis_path = input_path.with_suffix("").with_name(input_path.stem + "_analysis.json")

    if not analysis_path.exists():
        raise FileNotFoundError(
            f"Análisis no encontrado: {analysis_path}. "
            f"Correr antes xlsx_sheet_analyze.py."
        )

    with analysis_path.open(encoding="utf-8") as f:
        analysis = json.load(f)

    repr_choice = analysis["representation"]

    if repr_choice == "markdown_table":
        body = convert_markdown_table(input_path)
    elif repr_choice == "html_table":
        body = convert_html_table(input_path)
    elif repr_choice == "markdown_text":
        body = convert_markdown_text(input_path)
    elif repr_choice == "empty":
        body = "<!-- hoja vacía, omitida -->"
    else:
        raise ValueError(f"Representación desconocida: {repr_choice}")

    fm = _frontmatter(analysis, input_path)
    full = fm + body + "\n"

    if output_path is None:
        output_path = input_path.with_suffix(".md")

    with output_path.open("w", encoding="utf-8") as f:
        f.write(full)

    return {
        "input": str(input_path),
        "analysis": str(analysis_path),
        "output": str(output_path),
        "representation": repr_choice,
        "bytes_written": len(full.encode("utf-8")),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    parser.add_argument("input", help="XLSX single-tab.")
    parser.add_argument("--analysis", default=None, help="Path al JSON de análisis. Default: {input}_analysis.json.")
    parser.add_argument("--output", default=None, help="Path al MD/HTML de salida. Default: {input}.md.")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        sys.stderr.write(f"Input no existe: {input_path}\n")
        return 2

    analysis_path = Path(args.analysis) if args.analysis else None
    output_path = Path(args.output) if args.output else None

    result = convert(input_path, analysis_path, output_path)
    print(
        f"OK: repr={result['representation']} "
        f"bytes={result['bytes_written']} → {result['output']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
