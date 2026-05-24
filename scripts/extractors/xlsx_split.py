#!/usr/bin/env python3
"""
xlsx_split.py — Paso 1.0.b.1 del workflow.

Divide un archivo XLSX/XLSM multi-pestaña en varios archivos XLSX single-tab,
preservando metadata del origen para trazabilidad downstream.

Por qué este paso:
    Excel multi-pestaña no es procesable directamente porque cada hoja puede
    tener una representación óptima distinta (markdown table, HTML table, o texto).
    Separar a single-tab permite decidir y convertir cada hoja independientemente.

Antes del split, todas las celdas se materializan a valores (sin fórmulas) usando
los valores cacheados del workbook original. Esto evita referencias cross-sheet
rotas al eliminar hojas hermanas.

Uso:
    python3 xlsx_split.py INPUT.xlsx --output-dir DIR

Outputs (en DIR):
    {stem}_sheet{NN}_{slug}.xlsx           — un XLSX por hoja (solo valores)
    {stem}_split_manifest.json             — metadata: hojas, índices, paths, conteos

Convención de naming:
    NN: índice 1-based zero-padded (01, 02, ...).
    slug: nombre de hoja saneado a [a-z0-9_], cortado a 40 chars.

Determinístico. No LLM. Idempotente (sobrescribe outputs).

Dependencias: openpyxl
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write(
        "openpyxl no instalado. Instalar con:\n"
        "  pip install --break-system-packages openpyxl\n"
    )
    sys.exit(1)


def slugify(name: str, max_len: int = 40) -> str:
    """Sanitizar nombre de hoja para usar como parte de un filename."""
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    if not s:
        s = "sheet"
    return s[:max_len]


def _cell_has_formula(cell) -> bool:
    value = cell.value
    return isinstance(value, str) and value.startswith("=")


def _formula_references_other_sheet(formula: str, sheet_name: str) -> bool:
    if not formula.startswith("="):
        return False
    for match in re.finditer(r"'([^']+)'!", formula):
        if match.group(1) != sheet_name:
            return True
    for match in re.finditer(r"(?<![A-Za-z0-9_'])([A-Za-z_][A-Za-z0-9_. ]*)!", formula):
        if match.group(1) != sheet_name:
            return True
    return False


def materialize_workbook_values(input_path: Path) -> tuple[Path, dict[str, list[str]]]:
    """Replace formulas with cached/evaluated values before splitting."""
    wb_values = load_workbook(filename=str(input_path), data_only=True)
    wb = load_workbook(filename=str(input_path), data_only=False)
    sheet_warnings: dict[str, list[str]] = {}

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_values = wb_values[sheet_name]
            warnings: list[str] = []
            had_cross_sheet = False
            missing_cached = False

            for row in ws.iter_rows():
                for cell in row:
                    cached = ws_values.cell(cell.row, cell.column).value
                    if _cell_has_formula(cell):
                        formula = str(cell.value)
                        if _formula_references_other_sheet(formula, sheet_name):
                            had_cross_sheet = True
                        if cached is None or cached == "":
                            missing_cached = True
                        cell.value = cached
                    elif cached is not None:
                        cell.value = cached

            if had_cross_sheet:
                warnings.append("cross_sheet_formulas_materialized_to_values")
            if missing_cached:
                warnings.append("formula_cells_missing_cached_values")
            if warnings:
                sheet_warnings[sheet_name] = warnings

        tmp = tempfile.NamedTemporaryFile(
            prefix=f"{input_path.stem}_values_",
            suffix=input_path.suffix.lower(),
            delete=False,
        )
        tmp_path = Path(tmp.name)
        tmp.close()
        wb.save(str(tmp_path))
        return tmp_path, sheet_warnings
    finally:
        del wb_values
        del wb


def split_xlsx(input_path: Path, output_dir: Path) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)

    values_path, sheet_warnings = materialize_workbook_values(input_path)
    stem = input_path.stem
    sheets_meta = []

    try:
        wb_values = load_workbook(filename=str(values_path), data_only=True)
        try:
            for idx, sheet_name in enumerate(wb_values.sheetnames, start=1):
                ws_values = wb_values[sheet_name]

                non_empty_cells = 0
                for row in ws_values.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None and cell != "":
                            non_empty_cells += 1

                warnings = list(sheet_warnings.get(sheet_name, []))
                if non_empty_cells == 0:
                    warnings.append("sheet_appears_empty_after_value_materialization")

                slug = slugify(sheet_name)
                out_name = f"{stem}_sheet{idx:02d}_{slug}.xlsx"
                out_path = output_dir / out_name

                wb_single = load_workbook(filename=str(values_path), data_only=False)
                try:
                    for other_name in list(wb_single.sheetnames):
                        if other_name != sheet_name:
                            del wb_single[other_name]
                    wb_single.save(str(out_path))
                finally:
                    del wb_single

                sheets_meta.append({
                    "index": idx,
                    "original_name": sheet_name,
                    "slug": slug,
                    "output_path": str(out_path),
                    "non_empty_cells": non_empty_cells,
                    "has_formulas": False,
                    "values_materialized": True,
                    "warnings": warnings,
                })
        finally:
            del wb_values
    finally:
        values_path.unlink(missing_ok=True)

    manifest = {
        "source": str(input_path),
        "stem": stem,
        "total_sheets": len(sheets_meta),
        "output_dir": str(output_dir),
        "values_materialized": True,
        "sheets": sheets_meta,
    }

    manifest_path = output_dir / f"{stem}_split_manifest.json"
    with manifest_path.open("w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    return manifest


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    parser.add_argument("input", help="Archivo XLSX/XLSM de entrada (multi-pestaña).")
    parser.add_argument(
        "--output-dir", required=True,
        help="Directorio donde escribir los XLSX single-tab + manifest.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        sys.stderr.write(f"Input no existe: {input_path}\n")
        return 2
    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        sys.stderr.write(
            f"Formato no soportado: {input_path.suffix}. Esperado .xlsx o .xlsm\n"
        )
        return 2

    output_dir = Path(args.output_dir)
    manifest = split_xlsx(input_path, output_dir)

    print(f"OK: {manifest['stem']} → {manifest['total_sheets']} hojas en {output_dir}")
    for s in manifest["sheets"]:
        warn_str = f" ⚠ {','.join(s['warnings'])}" if s["warnings"] else ""
        print(
            f"  [{s['index']:02d}] {s['original_name']!r:40s} "
            f"({s['non_empty_cells']} celdas){warn_str}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
