#!/usr/bin/env python3
"""
xlsx_sheet_analyze.py — Paso 1.0.b.2 del workflow.

Dado un XLSX single-tab (output de xlsx_split.py), analiza la hoja y decide
qué representación es óptima para que un LLM downstream consuma sus datos.

Decisión:
    - "markdown_table": tabla rectangular limpia, sin merged cells, headers en fila 1,
                        ratio de columnas/filas razonable.
    - "html_table":     hay merged cells (en headers o datos), o headers jerárquicos,
                        o estructura no rectangular. HTML preserva colspan/rowspan/scope.
    - "markdown_text":  contenido dominante es texto largo (no tabular), pocas filas
                        con celdas grandes. Típicamente "instrucciones a postores"
                        pegadas en una hoja.
    - "empty":          hoja vacía o sin contenido útil.

Uso:
    python3 xlsx_sheet_analyze.py INPUT_SINGLE_TAB.xlsx [--output-json PATH]

Outputs:
    - stdout: linea resumen "REPR={...} merged={n} rows={r} cols={c}"
    - JSON con detalles si --output-json (default: junto al input con sufijo _analysis.json)

Reglas de decisión (umbrales configurables vía CLI):
    --text-cell-threshold       length mínimo para considerar una celda "texto largo"
                                (default: 200 chars)
    --text-cell-ratio           fracción mínima de celdas no vacías que son "texto largo"
                                para clasificar como markdown_text (default: 0.40)
    --header-merged-triggers-html  si hay merged cells en las primeras 3 filas → html
                                (default: True)
    --any-merged-data-triggers-html  si hay merged cells en filas de datos → html
                                (default: True)

Determinístico. No LLM. Idempotente.

Dependencias: openpyxl
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except ImportError:
    sys.stderr.write(
        "openpyxl no instalado. Instalar con:\n"
        "  pip install --break-system-packages openpyxl\n"
    )
    sys.exit(1)


def analyze_sheet(
    input_path: Path,
    text_cell_threshold: int = 200,
    text_cell_ratio: float = 0.40,
    header_merged_triggers_html: bool = True,
    any_merged_data_triggers_html: bool = True,
) -> dict:
    wb = load_workbook(filename=str(input_path), data_only=True)
    if len(wb.sheetnames) != 1:
        raise ValueError(
            f"Esperado XLSX single-tab; este archivo tiene {len(wb.sheetnames)} hojas. "
            f"Pasarlo primero por xlsx_split.py."
        )
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Merged cells
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    merged_count = len(merged_ranges)

    # ¿Hay merged cells en las primeras 3 filas? → jerarquía de headers.
    merged_in_header_rows = 0
    merged_in_data_rows = 0
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col_r, max_row_r = range_boundaries(str(rng))
        if min_row <= 3:
            merged_in_header_rows += 1
        else:
            merged_in_data_rows += 1

    # Conteo de celdas no vacías y métricas de "texto largo"
    non_empty_cells = 0
    long_text_cells = 0
    total_text_chars = 0
    cells_with_text = 0
    cells_with_number = 0

    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None or cell == "":
                continue
            non_empty_cells += 1
            if isinstance(cell, str):
                cells_with_text += 1
                total_text_chars += len(cell)
                if len(cell) >= text_cell_threshold:
                    long_text_cells += 1
            elif isinstance(cell, (int, float)):
                cells_with_number += 1
            # bool, datetime, etc. cuentan como non_empty pero no caen en text/number

    # Ratio de "texto largo"
    if non_empty_cells > 0:
        long_text_ratio = long_text_cells / non_empty_cells
    else:
        long_text_ratio = 0.0

    # Aspect ratio aproximado: filas vs columnas
    if max_col > 0:
        aspect_ratio = max_row / max_col
    else:
        aspect_ratio = 0.0

    # ----- DECISIÓN -----
    decision = None
    reasons = []

    if non_empty_cells == 0:
        decision = "empty"
        reasons.append("no_non_empty_cells")

    elif long_text_ratio >= text_cell_ratio and max_col <= 3:
        # Hoja narrativa: muchas celdas con texto largo, pocas columnas
        decision = "markdown_text"
        reasons.append(
            f"long_text_ratio={long_text_ratio:.2f} ≥ {text_cell_ratio} "
            f"and max_col={max_col} ≤ 3"
        )

    elif header_merged_triggers_html and merged_in_header_rows > 0:
        decision = "html_table"
        reasons.append(f"merged_in_header_rows={merged_in_header_rows} > 0")

    elif any_merged_data_triggers_html and merged_in_data_rows > 0:
        decision = "html_table"
        reasons.append(f"merged_in_data_rows={merged_in_data_rows} > 0")

    else:
        decision = "markdown_table"
        reasons.append("rectangular_no_merged_no_long_text")

    analysis = {
        "input": str(input_path),
        "sheet_name": sheet_name,
        "max_row": max_row,
        "max_col": max_col,
        "non_empty_cells": non_empty_cells,
        "cells_with_text": cells_with_text,
        "cells_with_number": cells_with_number,
        "long_text_cells": long_text_cells,
        "long_text_ratio": round(long_text_ratio, 4),
        "total_text_chars": total_text_chars,
        "aspect_ratio_rows_over_cols": round(aspect_ratio, 4),
        "merged_count": merged_count,
        "merged_in_header_rows": merged_in_header_rows,
        "merged_in_data_rows": merged_in_data_rows,
        "merged_ranges": merged_ranges,
        "thresholds": {
            "text_cell_threshold": text_cell_threshold,
            "text_cell_ratio": text_cell_ratio,
            "header_merged_triggers_html": header_merged_triggers_html,
            "any_merged_data_triggers_html": any_merged_data_triggers_html,
        },
        "representation": decision,
        "reasons": reasons,
    }
    return analysis


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    parser.add_argument("input", help="XLSX single-tab a analizar.")
    parser.add_argument("--output-json", default=None, help="Path del JSON de salida.")
    parser.add_argument("--text-cell-threshold", type=int, default=200)
    parser.add_argument("--text-cell-ratio", type=float, default=0.40)
    parser.add_argument("--no-header-merged-triggers-html", action="store_true")
    parser.add_argument("--no-any-merged-data-triggers-html", action="store_true")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        sys.stderr.write(f"Input no existe: {input_path}\n")
        return 2

    analysis = analyze_sheet(
        input_path,
        text_cell_threshold=args.text_cell_threshold,
        text_cell_ratio=args.text_cell_ratio,
        header_merged_triggers_html=not args.no_header_merged_triggers_html,
        any_merged_data_triggers_html=not args.no_any_merged_data_triggers_html,
    )

    if args.output_json:
        out_path = Path(args.output_json)
    else:
        out_path = input_path.with_suffix("").with_name(input_path.stem + "_analysis.json")

    with out_path.open("w", encoding="utf-8") as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)

    print(
        f"REPR={analysis['representation']} "
        f"merged={analysis['merged_count']} "
        f"rows={analysis['max_row']} cols={analysis['max_col']} "
        f"non_empty={analysis['non_empty_cells']} "
        f"long_text_ratio={analysis['long_text_ratio']:.2f}"
    )
    print(f"  → {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
