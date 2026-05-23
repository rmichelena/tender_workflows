#!/usr/bin/env python3
"""
Paso 7.1 — Consolidación Final
Lee todos los resultados de búsqueda, BOM búsqueda, BOM exploded, y specs.
Produce consolidado.json + TSV + MD + XLSX.
"""
import json, os, sys, csv
sys.path.insert(0, '/opt/data/home/.local/lib/python3.13/site-packages')
from collections import Counter, OrderedDict
from datetime import date

BASE = "/opt/data/workspace/tender_procurement/proyecto"
BOM_BUSQUEDA = f"{BASE}/artifacts/step_4_busqueda/BOM_busqueda.json"
BOM_EXPLODED = f"{BASE}/artifacts/step_2_bom/BOM_exploded_consolidado.json"
RESULTADOS_DIR = f"{BASE}/artifacts/step_6_resultados/items"
SPECS_DIR = f"{BASE}/artifacts/step_3_specs"
OUTPUT_DIR = f"{BASE}/outputs"

os.makedirs(OUTPUT_DIR, exist_ok=True)

def load_json(path):
    with open(path, 'r') as f:
        return json.load(f)

def get_spec_clave(item_id):
    """Extract top 2-3 hard specs as spec_clave string."""
    spec_path = os.path.join(SPECS_DIR, f"ITEM-{item_id}_specs.json")
    if not os.path.exists(spec_path):
        return ""
    try:
        specs = load_json(spec_path)
        hard_reqs = [r for r in specs.get("requerimientos", []) if r.get("hard_soft") == "Hard"]
        # Take up to 3 most distinctive
        claves = []
        for r in hard_reqs[:5]:
            param = r.get("parametro", "")
            valor = r.get("valor", "")
            if param and valor:
                claves.append(f"{param}: {valor}")
            elif param:
                claves.append(param)
            if len(claves) >= 3:
                break
        return "; ".join(claves)[:120]
    except Exception as e:
        return ""

def main():
    print("=== Paso 7.1 — Consolidación Final ===")
    
    # Load inputs
    bom_busqueda = load_json(BOM_BUSQUEDA)
    bom_exploded = load_json(BOM_EXPLODED)
    
    bienes_ids = set(i["id"] for i in bom_busqueda["items"])
    servicios = [i for i in bom_exploded["items"] if i["grupo"] == "Servicios"]
    
    print(f"BOM búsqueda: {len(bienes_ids)} bienes")
    print(f"BOM exploded: {len(bom_exploded['items'])} items ({len(servicios)} servicios)")
    
    # Build item lookup from BOM exploded
    item_lookup = {}
    for item in bom_exploded["items"]:
        item_lookup[item["id"]] = item
    
    # Build group ordering
    grupo_order = {"VSAT/HUB": 1, "Radio Comms": 2, "Energía": 3, "Infraestructura": 4, "Servicios": 5}
    
    # Process items
    filas = []
    stats = Counter()
    
    # Sort bienes by grupo then id
    bienes_sorted = sorted(bom_busqueda["items"], key=lambda x: (grupo_order.get(x["grupo"], 99), x["id"]))
    
    for bien in bienes_sorted:
        item_id = bien["id"]
        grupo = bien["grupo"]
        desc = bien.get("descripcion_limpia", item_lookup.get(item_id, {}).get("descripcion", ""))
        spec_clave = get_spec_clave(item_id)
        
        # Try to read resultado
        result_path = os.path.join(RESULTADOS_DIR, f"ITEM-{item_id}_resultado.json")
        ruta_resultado = f"artifacts/step_6_resultados/items/ITEM-{item_id}_resultado.json"
        
        if not os.path.exists(result_path):
            # No result file — SIN_CANDIDATO
            filas.append({
                "item_id": item_id,
                "nombre": desc[:120],
                "grupo": grupo,
                "tipo": "BIEN",
                "spec_clave": spec_clave,
                "candidato_num": 0,
                "marca": "",
                "modelo": "",
                "part_number": "",
                "estado": "SIN_CANDIDATO",
                "url_fabricante": "",
                "url_datasheet": "",
                "resumen_cumplimiento": "",
                "notas": "Sin archivo de resultado disponible",
                "ruta_matriz": "",
                "ruta_resultado": ruta_resultado
            })
            stats["sin_candidato"] += 1
            continue
        
        resultado = load_json(result_path)
        candidatos = resultado.get("candidatos", [])
        
        # Filter only VALIDO or CONDICIONADO
        valid_cands = [c for c in candidatos if c.get("clasificacion") in ("VALIDO", "CONDICIONADO")]
        
        if not valid_cands:
            filas.append({
                "item_id": item_id,
                "nombre": desc[:120],
                "grupo": grupo,
                "tipo": "BIEN",
                "spec_clave": spec_clave,
                "candidato_num": 0,
                "marca": "",
                "modelo": "",
                "part_number": "",
                "estado": "SIN_CANDIDATO",
                "url_fabricante": "",
                "url_datasheet": "",
                "resumen_cumplimiento": "",
                "notas": resultado.get("diagnostico", "Sin candidatos válidos tras búsqueda"),
                "ruta_matriz": "",
                "ruta_resultado": ruta_resultado
            })
            stats["sin_candidato"] += 1
        else:
            for idx, cand in enumerate(valid_cands, 1):
                estado = cand.get("clasificacion", "DESCARTADO")
                if estado == "VALIDO":
                    stats["validos"] += 1
                else:
                    stats["condicionados"] += 1
                
                filas.append({
                    "item_id": item_id,
                    "nombre": desc[:120],
                    "grupo": grupo,
                    "tipo": "BIEN",
                    "spec_clave": spec_clave,
                    "candidato_num": idx,
                    "marca": cand.get("marca", ""),
                    "modelo": cand.get("modelo", ""),
                    "part_number": cand.get("part_number", ""),
                    "estado": estado,
                    "url_fabricante": cand.get("url_fabricante", cand.get("url_fuente", "")),
                    "url_datasheet": cand.get("url_datasheet", ""),
                    "resumen_cumplimiento": cand.get("resumen_cumplimiento", ""),
                    "notas": cand.get("notas", ""),
                    "ruta_matriz": cand.get("ruta_matriz", ""),
                    "ruta_resultado": ruta_resultado
                })
            
            if any(c.get("clasificacion") == "VALIDO" for c in valid_cands):
                stats["items_resueltos"] += 1
            else:
                stats["items_solo_condicionados"] += 1
    
    # Add services
    for svc in servicios:
        item_id = svc["id"]
        filas.append({
            "item_id": item_id,
            "nombre": svc["descripcion"][:120],
            "grupo": "Servicios",
            "tipo": "SERVICIO",
            "spec_clave": "",
            "candidato_num": 0,
            "marca": "",
            "modelo": "",
            "part_number": "",
            "estado": "SERVICIO",
            "url_fabricante": "",
            "url_datasheet": "",
            "resumen_cumplimiento": "",
            "notas": "Servicio — no requiere búsqueda de producto",
            "ruta_matriz": "",
            "ruta_resultado": ""
        })
        stats["servicios"] += 1
    
    # Sort filas: by grupo order, then item_id, then candidato_num
    filas.sort(key=lambda x: (grupo_order.get(x["grupo"], 99), x["item_id"], x["candidato_num"]))
    
    # Compute totals
    items_bienes = len(bienes_ids)
    items_servicios = len(servicios)
    items_total = items_bienes + items_servicios
    candidatos_totales = stats["validos"] + stats["condicionados"]
    
    # Count unique items with candidatos
    items_con_cand = len(set(f["item_id"] for f in filas if f["candidato_num"] > 0))
    items_sin_cand = items_bienes - items_con_cand
    
    totales = {
        "items_total": items_total,
        "items_bienes": items_bienes,
        "items_servicios": items_servicios,
        "items_con_candidatos": items_con_cand,
        "items_sin_candidato": items_sin_cand,
        "items_resueltos_con_valido": stats.get("items_resueltos", 0),
        "items_solo_condicionados": stats.get("items_solo_condicionados", 0),
        "candidatos_totales": candidatos_totales,
        "candidatos_validos": stats.get("validos", 0),
        "candidatos_condicionados": stats.get("condicionados", 0)
    }
    
    # Build consolidado JSON
    consolidado = {
        "tipo": "CONSOLIDADO_FINAL",
        "proyecto": "ICAO-00068",
        "fecha_generacion": str(date.today()),
        "totales": totales,
        "filas": filas
    }
    
    # Write JSON
    json_path = f"{OUTPUT_DIR}/consolidado.json"
    with open(json_path, 'w') as f:
        json.dump(consolidado, f, indent=2, ensure_ascii=False)
    print(f"JSON: {json_path} ({len(filas)} filas)")
    
    # Write TSV
    tsv_path = f"{OUTPUT_DIR}/consolidado.tsv"
    fieldnames = ["item_id", "nombre", "grupo", "tipo", "spec_clave", "candidato_num",
                  "marca", "modelo", "part_number", "estado", "url_fabricante", "url_datasheet",
                  "resumen_cumplimiento", "notas", "ruta_matriz", "ruta_resultado"]
    with open(tsv_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter='\t', extrasaction='ignore')
        writer.writeheader()
        writer.writerows(filas)
    print(f"TSV: {tsv_path}")
    
    # Write MD
    md_path = f"{OUTPUT_DIR}/consolidado.md"
    with open(md_path, 'w') as f:
        f.write(f"# Consolidado Final — ICAO-00068\n\n")
        f.write(f"**Fecha**: {date.today()}\n\n")
        f.write(f"## Resumen\n\n")
        f.write(f"- **Items totales**: {items_total} ({items_bienes} bienes + {items_servicios} servicios)\n")
        f.write(f"- **Con candidatos**: {items_con_cand}\n")
        f.write(f"- **Sin candidato**: {items_sin_cand}\n")
        f.write(f"- **Candidatos totales**: {candidatos_totales} ({stats.get('validos',0)} válidos + {stats.get('condicionados',0)} condicionados)\n\n")
        
        # Table by group
        f.write(f"## Por Grupo\n\n")
        grp_counts = Counter()
        grp_cands = Counter()
        for fila in filas:
            if fila["tipo"] == "BIEN":
                grp_counts[fila["grupo"]] += 1
                if fila["candidato_num"] > 0:
                    grp_cands[fila["grupo"]] += 1
        
        for grp in sorted(grp_counts.keys(), key=lambda x: grupo_order.get(x, 99)):
            f.write(f"- **{grp}**: {grp_counts[grp]} items, {grp_cands[grp]} con candidatos\n")
        
        f.write(f"\n## Detalle de Candidatos\n\n")
        
        current_item = None
        for fila in filas:
            if fila["item_id"] != current_item:
                current_item = fila["item_id"]
                f.write(f"\n### {fila['item_id']} — {fila['nombre'][:80]}\n")
                f.write(f"- Grupo: {fila['grupo']} | Tipo: {fila['tipo']}\n")
                if fila["spec_clave"]:
                    f.write(f"- Specs clave: {fila['spec_clave']}\n")
            
            if fila["candidato_num"] > 0:
                f.write(f"  - **Candidato {fila['candidato_num']}**: {fila['marca']} {fila['modelo']}")
                if fila["part_number"]:
                    f.write(f" ({fila['part_number']})")
                f.write(f" — **{fila['estado']}**\n")
                if fila["url_fabricante"]:
                    f.write(f"    - URL: {fila['url_fabricante']}\n")
                if fila["notas"]:
                    f.write(f"    - Notas: {fila['notas'][:200]}\n")
            elif fila["estado"] == "SIN_CANDIDATO":
                f.write(f"  - ❌ **SIN CANDIDATO**: {fila['notas'][:200]}\n")
            elif fila["estado"] == "SERVICIO":
                f.write(f"  - 📋 Servicio\n")
    
    print(f"MD: {md_path}")
    
    # Try XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        
        # Hoja Consolidado
        ws = wb.active
        ws.title = "Consolidado"
        
        # Headers
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for col, h in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        estado_fills = {
            "VALIDO": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "CONDICIONADO": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "SIN_CANDIDATO": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "SERVICIO": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        }
        
        for row_idx, fila in enumerate(filas, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(fila.get(field, "")))
            # Color estado
            estado = fila.get("estado", "")
            if estado in estado_fills:
                for col_idx in range(1, len(fieldnames)+1):
                    ws.cell(row=row_idx, column=col_idx).fill = estado_fills[estado]
        
        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Hoja Resumen
        ws2 = wb.create_sheet("Resumen")
        ws2.cell(1, 1, "Resumen Consolidado ICAO-00068").font = Font(bold=True, size=14)
        ws2.cell(3, 1, "Métrica").font = Font(bold=True)
        ws2.cell(3, 2, "Valor").font = Font(bold=True)
        for idx, (k, v) in enumerate(totales.items(), 4):
            ws2.cell(idx, 1, k.replace("_", " ").title())
            ws2.cell(idx, 2, v)
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 15
        
        xlsx_path = f"{OUTPUT_DIR}/consolidado.xlsx"
        wb.save(xlsx_path)
        print(f"XLSX: {xlsx_path}")
    except ImportError:
        print("openpyxl no disponible, saltando XLSX")
    
    # Print summary
    print(f"\n=== RESUMEN ===")
    print(f"{items_total} ítems | {items_con_cand} con candidatos | "
          f"{stats.get('items_solo_condicionados',0)} solo condicionados | "
          f"{items_sin_cand} sin candidato | "
          f"{candidatos_totales} candidatos totales | "
          f"{items_servicios} servicios")
    print("OK")

if __name__ == "__main__":
    main()
